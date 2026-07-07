"""
Full calibration pipeline:
  1. Run EPS.mdl in Vensim DSS (BAU, no .cin) and export BAU cost-per-mile
     and BAU new-vehicles-demanded to a .tsv file.
  2. Parse the .tsv export and update calibration_parameters.csv with fresh
     cost_per_mile and new_vehicle_demand values for the TARGET_YEARS.
  3. Run transport_calibration.py to recompute shareweights.

Usage:
    py run_calibration_pipeline.py          (runs steps 1-3)
    py run_calibration_pipeline.py --skip-vensim   (skips step 1, uses existing .tsv)
"""

import os
import sys
import csv
import re
import subprocess

# =========================
# Paths
# =========================
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EPS_DIR    = os.path.abspath(os.path.join(BASE_DIR, "..", "..", ".."))
MDL_FILE   = os.path.join(EPS_DIR, "EPS.mdl")
_VENSIM_CANDIDATES = [
    r"C:\Program Files\Vensim DSS x64\vendss64.exe",
    r"C:\Program Files\Vensim\vendss64.exe",
    r"C:\Program Files (x86)\Vensim\vendss.exe",
]
VENSIM_EXE = next((p for p in _VENSIM_CANDIDATES if os.path.exists(p)), _VENSIM_CANDIDATES[0])

EXPORT_LST = os.path.join(BASE_DIR, "calibration_export_vars.lst")
EXPORT_TSV = os.path.join(BASE_DIR, "bau_export.tsv")
CMD_FILE   = os.path.join(BASE_DIR, "run_and_export.cmd")

CALIB_DIR  = os.path.join(BASE_DIR, "calibration")
CALIB_CSV  = os.path.join(CALIB_DIR, "calibration_parameters.csv")
CALIB_PY   = os.path.join(CALIB_DIR, "transport_calibration.py")

RUN_NAME   = "BAU_calibration"

# Which years to update in calibration_parameters.csv
TARGET_YEARS = [2024, 2025]

# Model time range
INITIAL_YEAR = 2024
FINAL_YEAR   = 2050


# =========================
# Step 1: Run Vensim DSS
# =========================
def run_vensim():
    if not os.path.exists(VENSIM_EXE):
        print(f"ERROR: Vensim DSS not found at {VENSIM_EXE}")
        sys.exit(1)
    if not os.path.exists(MDL_FILE):
        print(f"ERROR: Model file not found at {MDL_FILE}")
        sys.exit(1)

    # Write the command script (always regenerate to ensure paths are correct)
    vdfx_path = os.path.join(EPS_DIR, f"{RUN_NAME}.vdfx")
    cmd_content = (
        f'SPECIAL>LOADMODEL|"{MDL_FILE}"\n'
        f'\n'
        f'SIMULATE>SAVELIST|{EXPORT_LST}\n'
        f'SIMULATE>RUNNAME|{RUN_NAME}\n'
        f'SIMULATE>READCIN|\n'
        f'MENU>RUN|O\n'
        f'MENU>VDF2TAB|{vdfx_path}|{EXPORT_TSV}|{EXPORT_LST}|+!||{INITIAL_YEAR}|{FINAL_YEAR}|:{RUN_NAME}\n'
        f'FILE>DELETE|{vdfx_path}\n'
        f'\n'
        f'SIMULATE>SAVELIST|\n'
        f'MENU>EXIT\n'
    )
    with open(CMD_FILE, "w") as f:
        f.write(cmd_content)

    print(f"Running Vensim DSS on {MDL_FILE}...")
    result = subprocess.run(
        [VENSIM_EXE, "", CMD_FILE],
        cwd=EPS_DIR,
        capture_output=True,
        text=True,
        timeout=600,
    )
    if not os.path.exists(EXPORT_TSV):
        print("ERROR: Vensim did not produce the export file.")
        print("stdout:", result.stdout)
        print("stderr:", result.stderr)
        sys.exit(1)

    print(f"OK: Vensim export written to {EXPORT_TSV}")


# =========================
# Step 2: Parse TSV and update calibration_parameters.csv
# =========================
def parse_vensim_variable(name):
    """Parse 'BAU New Vehicle Cost per Mile[LDVs,passenger,battery electric vehicle]'
    into (variable_name, vehicle_type, cargo_type, technology)."""
    m = re.match(r"^(.+?)\[(.+)\]$", name.strip())
    if not m:
        return name.strip(), None, None, None
    var = m.group(1).strip()
    subs = [s.strip() for s in m.group(2).split(",")]
    if len(subs) == 3:
        return var, subs[0], subs[1], subs[2]
    elif len(subs) == 2:
        return var, subs[0], subs[1], None
    return var, None, None, None


VT_MAP = {
    ("LDVs", "passenger"): "passenger LDV",
    ("LDVs", "freight"):   "freight LDV",
    ("HDVs", "passenger"): "passenger HDV",
    ("HDVs", "freight"):   "freight HDV",
}


def update_calibration_csv():
    if not os.path.exists(EXPORT_TSV):
        print(f"ERROR: Export TSV not found at {EXPORT_TSV}")
        sys.exit(1)

    # ---- Parse the .tsv ----
    years = list(range(INITIAL_YEAR, FINAL_YEAR + 1))
    cost_data = {}   # (vt_label, tech, year) -> cost_per_mile
    demand_data = {} # (vt_label, year) -> new_vehicle_demand

    with open(EXPORT_TSV, "r") as f:
        reader = csv.reader(f, delimiter="\t")
        for row in reader:
            if len(row) < 3:
                continue
            var_full = row[0]
            # row[1] is run name; row[2:] are values for each year
            values = row[2:]

            var, vt, ct, tech = parse_vensim_variable(var_full)

            if vt and ct:
                vt_label = VT_MAP.get((vt, ct))
                if not vt_label:
                    continue
            else:
                continue

            if "Cost per Mile" in var and tech:
                for i, y in enumerate(years):
                    if i < len(values):
                        try:
                            cost_data[(vt_label, tech, y)] = float(values[i])
                        except ValueError:
                            pass
            elif "Vehicles Demanded" in var:
                for i, y in enumerate(years):
                    if i < len(values):
                        try:
                            demand_data[(vt_label, y)] = float(values[i])
                        except ValueError:
                            pass

    # ---- Read existing calibration_parameters.csv ----
    with open(CALIB_CSV, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)

    # ---- Update cost_per_mile and new_vehicle_demand for TARGET_YEARS ----
    updated = 0
    for row in rows:
        yr = int(row["year"])
        if yr not in TARGET_YEARS:
            continue

        vt = row["vehicle type"].strip()
        tech = row["tech"].strip()

        # Map vehicle_type from CSV to our label
        vt_lower = vt.lower()
        if "passenger" in vt_lower and "hdv" in vt_lower:
            vt_label = "passenger HDV"
        elif "passenger" in vt_lower and "ldv" in vt_lower:
            vt_label = "passenger LDV"
        elif "freight" in vt_lower and "hdv" in vt_lower:
            vt_label = "freight HDV"
        elif "freight" in vt_lower and "ldv" in vt_lower:
            vt_label = "freight LDV"
        else:
            continue

        key_cost = (vt_label, tech, yr)
        key_demand = (vt_label, yr)

        if key_cost in cost_data:
            old = row["cost_per_mile"]
            row["cost_per_mile"] = str(cost_data[key_cost])
            if old != row["cost_per_mile"]:
                updated += 1

        if key_demand in demand_data:
            row["new_vehicle_demand"] = str(demand_data[key_demand])

    # ---- Write updated CSV ----
    with open(CALIB_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"OK: Updated {updated} cost_per_mile values in {CALIB_CSV}")


# =========================
# Step 3: Run transport_calibration.py
# =========================
def run_calibration():
    print(f"Running {CALIB_PY}...")
    result = subprocess.run(
        [sys.executable, CALIB_PY],
        cwd=CALIB_DIR,
        capture_output=True,
        text=True,
    )
    print(result.stdout)
    if result.stderr:
        # Filter out expected warnings
        for line in result.stderr.splitlines():
            if "RuntimeWarning" not in line and "arraylike" not in line:
                print(line, file=sys.stderr)
    if result.returncode != 0:
        print(f"WARNING: calibration script exited with code {result.returncode}")


# =========================
# Step 4: Update Transportation Technology Shareweights.xlsx
# =========================
XLSX_FILE = os.path.join(BASE_DIR, "Transportation Technology Shareweights.xlsx")

# Mapping from (section_label in shareweights_formatted.csv) to Data-sheet start row.
# Each block is 7 techs in order: BEV, NG, Gas, Diesel, PHEV, LPG, H2
SECTION_TO_ROW = {
    "passenger LDV": 10,
    "passenger HDV": 24,
    "freight LDV":   17,
    "freight HDV":   31,
}
TECH_ORDER = [
    "battery electric vehicle",
    "natural gas vehicle",
    "gasoline vehicle",
    "diesel vehicle",
    "plugin hybrid vehicle",
    "LPG vehicle",
    "hydrogen vehicle",
]

def update_xlsx():
    """Parse shareweights_formatted.csv and write 2024/2025 values into
    columns D/E of the Data sheet in Transportation Technology Shareweights.xlsx."""
    from openpyxl import load_workbook

    sw_file = os.path.join(CALIB_DIR, "shareweights_formatted.csv")
    if not os.path.exists(sw_file):
        print(f"ERROR: {sw_file} not found. Run calibration first.")
        return
    if not os.path.exists(XLSX_FILE):
        print(f"ERROR: {XLSX_FILE} not found.")
        return

    # ---- Parse shareweights_formatted.csv ----
    # Format: section header line, then header ",2024,2025", then 7 tech rows, blank line
    values = {}  # (section, tech) -> (val_2024, val_2025)
    current_section = None
    with open(sw_file, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            if not row:
                continue
            cell0 = row[0].strip()
            # Check if this is a section header (e.g. "passenger LDV")
            if cell0 in SECTION_TO_ROW:
                current_section = cell0
                continue
            # Blank first cell = year-header row or separator; don't reset section
            if not cell0:
                # Only reset section on truly blank lines (all cells empty)
                if all(not c.strip() for c in row):
                    current_section = None
                continue
            # This should be a tech row
            if current_section and cell0 in TECH_ORDER:
                v24 = float(row[1]) if len(row) > 1 and row[1].strip() else 0.0
                v25 = float(row[2]) if len(row) > 2 and row[2].strip() else 0.0
                values[(current_section, cell0)] = (v24, v25)

    # ---- Write to xlsx Data sheet ----
    wb = load_workbook(XLSX_FILE)
    ws = wb["Data"]

    updated = 0
    for section, start_row in SECTION_TO_ROW.items():
        for i, tech in enumerate(TECH_ORDER):
            data_row = start_row + i
            key = (section, tech)
            if key in values:
                v24, v25 = values[key]
                ws.cell(row=data_row, column=4, value=v24)  # col D = 2024
                ws.cell(row=data_row, column=5, value=v25)  # col E = 2025
                updated += 1

    wb.save(XLSX_FILE)
    print(f"OK: Updated {updated} (section, tech) pairs in {XLSX_FILE} Data sheet (cols D-E)")


# =========================
# Main
# =========================
def main():
    skip_vensim = "--skip-vensim" in sys.argv

    if not skip_vensim:
        run_vensim()
    else:
        print("Skipping Vensim run (--skip-vensim). Using existing .tsv export.")

    update_calibration_csv()
    run_calibration()
    update_xlsx()

    print("\nPipeline complete.")


if __name__ == "__main__":
    main()
