"""Export model-readable CSVs from the PEaWHRP source workbook (waste heat + process efficiency measures).

Reads the PEaWHRP-* tabs of "Process Efficiency and Waste Heat Recovery Parameters.xlsx"
(the full traceable pipeline: Kermeli Data -> economics -> dist engine -> export tabs).
Cached formula values are used, so save the workbook from Excel after editing it.
The old intermediate PEaWHRP.xlsx is superseded (kept as .superseded).

Run from this directory:  python ExportPEaWHRP.py

Outputs (read by EPS.mdl):
  PEaWHRP-WM.csv    Waste Heat Measure subscript elements (A2 down)
  PEaWHRP-PM.csv    Process Efficiency Measure subscript elements (A2 down)
  PEaWHRP-WMD.csv   long-form data, one row per (industry x waste heat measure);
                    cols A-J from the xlsx tab + K = Key (flattened subscript element name).
                    Capex (G) and Opex (H) are divided by 1e6 at export:
                    2012$/MMBtu-yr in the xlsx -> 2012$/(BTU/yr) in the CSV.
  PEaWHRP-PMD.csv   same for (industry x efficiency measure)

Notes:
  - Commas inside text columns (Measure, Source) are replaced with ';' because
    Vensim splits CSV rows naively on the delimiter.
  - Row order must remain industry-major in Industry Category declaration order
    with measures minor; the VECTOR ELM MAP remaps in EPS.mdl depend on it.
  - PEaWHRP-MHRP.csv, PEaWHRP-MCL-mult.csv, PEaWHRP-MCL-share.csv and
    PEaWHRP-SoCEMDSiaY.csv are maintained by hand, not by this script.
"""
import csv
import os

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Process Efficiency and Waste Heat Recovery Parameters.xlsx")

UNIT_SCALE_COLS = {6, 7}  # 0-based: G Capex, H Opex  (divide by 1e6)
TEXT_COLS = {0, 1, 2, 3, 9}  # A Industry, B Code, C Slot, D Measure, J Source

HEADER_OVERRIDES = {
    6: "Capex intensity (2012$/(BTU/yr) of savings capacity; xlsx value / 1e6)",
    7: "Opex intensity (2012$/yr per BTU/yr of savings capacity; xlsx value / 1e6)",
}


def clean_text(v):
    if v is None:
        return ""
    return str(v).replace(",", ";").strip()


def export_subscript(wb, tab, out_name, header):
    ws = wb[tab]
    rows = [[header]]
    for row in ws.iter_rows(values_only=True):
        if row[0] is not None:
            rows.append([clean_text(row[0])])
    path = os.path.join(HERE, out_name)
    with open(path, "w", newline="") as f:
        csv.writer(f).writerows(rows)
    print(f"wrote {out_name}: {len(rows) - 1} elements")


def export_data(wb, tab, out_name, key_prefix):
    ws = wb[tab]
    data = list(ws.iter_rows(values_only=True))
    header = list(data[0])
    for i, override in HEADER_OVERRIDES.items():
        header[i] = override
    header = [clean_text(h) for h in header] + ["Key"]
    out_rows = [header]
    for row in data[1:]:
        if row[0] is None:
            continue
        out = []
        for i in range(10):
            v = row[i]
            if i in TEXT_COLS:
                out.append(clean_text(v))
            elif i in UNIT_SCALE_COLS:
                out.append(repr(float(v or 0) / 1e6))
            else:
                out.append(repr(float(v or 0)))
        code = clean_text(row[1])
        slot = clean_text(row[2])
        out.append(f"{key_prefix} {code} X {slot}")
        out_rows.append(out)
    keys = [r[10] for r in out_rows[1:]]
    assert len(keys) == len(set(keys)), f"{out_name}: duplicate Key values"
    nonzero = sum(1 for r in out_rows[1:] if float(r[4]) != 0 or float(r[5]) != 0)
    assert nonzero >= 10, (
        f"{out_name}: only {nonzero} rows with nonzero savings - the workbook's cached "
        "formula values are missing (this happens after any programmatic save). "
        "Open the workbook in Excel, let it recalculate, save, and re-run this script.")
    path = os.path.join(HERE, out_name)
    with open(path, "w", newline="") as f:
        csv.writer(f).writerows(out_rows)
    print(f"wrote {out_name}: {len(out_rows) - 1} data rows")


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    export_subscript(wb, "PEaWHRP-WM", "PEaWHRP-WM.csv", "waste heat measures")
    export_subscript(wb, "PEaWHRP-PM", "PEaWHRP-PM.csv", "process efficiency measures")
    export_data(wb, "PEaWHRP-WMD", "PEaWHRP-WMD.csv", "whm")
    export_data(wb, "PEaWHRP-PMD", "PEaWHRP-PMD.csv", "pem")


if __name__ == "__main__":
    main()
